import os
import copy
import io
import re
import pickle
import fitz # PyMuPDF
import json
import mysql
import smtplib
import openpyxl
import xlrd
import datetime
import holidays
import random
import pandas as pd
import numpy as np

from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import formatting
from threading import Thread
from operator import itemgetter
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request, AuthorizedSession
from email.message import EmailMessage
from email.mime.text import MIMEText




# PATHS
db_config_path = "../config_files/db/config.json"
email_config_path = "../config_files/email_service/config.json"
google_config_path = "../config_files/google/google_userconfig.json"
logo_path = "../config_files/imgs/BusinessCat.png"
icon_path = "../config_files/imgs/Cat.ico"

# COLORS
default_background = "#ffffff"
color_light_orange = "#fff6e5"
color_green = "#009922"
color_yellow = "#ffe01a"
color_red = "#ff3333"
color_orange = "#e59345"
color_grey = "#e3e3e3"

# GOOGLE VARIABLES
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '../config_files/google/google_credentials.json'
creds = None
SCOPE = [
    'https://www.googleapis.com/auth/drive'
]


''' UTILITIES '''

def START_in_Thread(process):
    """prevent process to freeze the app upon launch"""
    Thread(target=process).start()

def check_paycheck_badges():

    done_paycheck = False
    done_badges = False

    if os.path.exists("BUSTE PAGA"):
        if len(os.listdir("BUSTE PAGA")) > 0:
            done_paycheck = True
    if os.path.exists("CARTELLINI"):
        if len(os.listdir("CARTELLINI")) > 0:
            done_badges = True

    check = (done_paycheck, done_badges)

    return check

def get_sheetnames_from_bytes(bytes_):
    """ open bytestream as openpyxl workbooks and return an array of worksheet names """
    wb = openpyxl.load_workbook(bytes_)
    sheetnames = wb.sheetnames
    return sheetnames


''' GOOGLE API METHODS '''

def authenticate(func):
    def auth_wrapper(*args, **kwargs):

        global creds
        global SCOPE

        # load token.pickle if present
        if os.path.exists('../config_files/google/token.pickle'):
            with open('../config_files/google/token.pickle', 'rb') as token:
                creds = pickle.load(token)

        if not creds or not creds.valid:
            # ? if token needs to be refreshed it will be refreshed
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            # ? otherwise authenticate
            else:
                flow = InstalledAppFlow.from_client_secrets_file(os.environ['GOOGLE_APPLICATION_CREDENTIALS'], SCOPE)
                creds = flow.run_local_server(port=0)

            with open('../config_files/google/token.pickle', 'wb') as token:
                pickle.dump(creds, token)

        # raise if user does not authenticate
        if not creds:
            raise Exception("You are not allowed to run this method >> Unauthorized")

        return func(*args, **kwargs)

    return auth_wrapper

@authenticate
def create_auth_session(credentials=creds):
    return AuthorizedSession(credentials)

@authenticate
def build_service(service, version="v3"):
    return build(service, version, credentials=creds)

def get_df_bytestream(ID):
    """
    get google sheet for comparison. return it as a bytestream
    """
    service = build_service("drive")
    conversion_table = service.about().get(fields="exportFormats").execute()
    request = service.files().export_media(fileId=ID,
                                           mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fh = io.BytesIO()

    # download file bytestream
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()

    return fh

def get_comparison_df(bytestream, month):
    """
    parse bytestream as a pandas dataframe
    """

    # parse bytestream into df
    df = pd.read_excel(bytestream, sheet_name=month.upper(), index_col=0)
    return df

def get_sheetlist():
    """
    get google sheet list in google drive
    """
    service = build_service("drive")
    results = service.files().list(pageSize=10, fields="nextPageToken, files(id, name, mimeType)", q="mimeType='application/vnd.google-apps.spreadsheet'").execute()
    items = results.get('files', [])
    items = sorted(items, key=itemgetter('name'))
    return items


''' MAILS '''

def load_email_server_config():
    try:
        with open(email_config_path, 'r') as f:
            config = json.load(f)
        return config
    except:
        raise Exception("Cannot find email server config file")

def connect_to_mail_server():
    config = load_email_server_config()
    smtp = smtplib.SMTP_SSL(config['server'], config['port'], timeout=10)
    smtp.ehlo()
    try:
        smtp.login(config['email'], config['password'])
        print(f"""CONNECTED SUCCESSFULLY:\n
        -> SERVER {config['server']}\n
        -> PORT {config['port']}\n
        -> USER {config['email']}\n
        -> PWD {"*"*len(str(config['password']))}
        """)
    except:
        raise Exception("Login al server non riuscito")

    return smtp


''' DB UTILS '''

def load_db_config():
    try:
        with open(db_config_path, "r") as f:
            config = json.load(f)
    except:
        raise Exception("Cannot find db config file")

    return config

def connect_to_db(config):

    db = mysql.connector.connect(
        host=config['host'],
        password=config['password'],
        username=config['username'],
        database=config['database'],
        port=config['port']
    )

    return db

def add_user(db, cursor, email, pwd, product_key, workstations):
    workstations = json.dumps(workstations)

    task = f"INSERT INTO users (email, pwd, product_key, workstations) VALUES ('{email}','{pwd}','{product_key}','{workstations}')"
    print(task)

    cursor.execute(task)
    print(f"-> user added")
    db.commit()
    return cursor

def check_registered(cursor, email):
    already_registered = False

    task = f"SELECT * FROM users WHERE email = '{email}'"
    cursor.execute(task)

    if cursor.fetchone():
        already_registered = True

    return already_registered


''' CLASSES '''

class PaycheckController():
    def __init__(self):
        """
        paychecks_to_check (str) -> path to multiple pages pdf containing all paychecks to check
        badges_to_check (str) -> path to folder containing all badges files as .pdf
        """
        """
        try:
            self.__validate_data(paychecks_to_check, badges_to_check)
        except Exception() as e:
            raise Exception(e)
        """

        self.badges_path = "" # path to CARTELLINI folder
        self.paychecks_to_check = "" # lul_controllo
        self.conversion_table_path = "../config_files/conversion_table.json"

        self.verify_filename = "Verifica.xlsx" #name of the output verification xlsx
        self.highlight_error = "FFFFFFFF"
        self.default_configuration = {
                "col_codes": {},
                "col_to_extract": [
                    "Z01100",
                    "Z01160",
                    "Z00246",
                    "Z00255",
                    "Z01138",
                    "000279",
                    "001008",
                    "ZP0160",
                    "ZP0162",
                    "ZPS000",
                    "F02701",
                    "000282",
                    "003955",
                    "Z05031",
                    "Z05075",
                    "ZP0001",
                    "ZP0003",
                    "ZP0030",
                    "003951",
                    "002099",
                    "002101",
                    "003802",
                    "002100",
                    "F09080",
                    "F09100",
                    "F09130",
                    "F09081",
                    "Z50022",
                    "Z50023",
                    "Z51000",
                    "Z51010",
                    "Z05004",
                    "Z05041",
                    "Z05065",
                    "Z05060",
                    "ZP0029",
                    "000085",
                    "000229",
                    "000086",
                    "ZP8134",
                    "ZP8138",
                    "ZP8130",
                    "003450",
                    "003411",
                    "000283",
                    "000031",
                    "quota t.f.r.",
                    "quota t.f.r. a fondi"
                ],
                "merging_columns": {
                    "Ferie/fest. pagate": [
                        "Z01100",
                        "Z01160",
                        "Z01138",
                        "Z00255",
                        "Z51010",
                        "Z00246",
                        "Z51000",
                        "000031"
                    ],
                    "T.F.R.": [
                        "quota t.f.r.",
                        "quota t.f.r. a fondi",
                        "ZP8134",
                        "003450",
                        "fondo t.f.r. al 31/12"
                    ],
                    "Cessione 1/5": [
                        "002101",
                        "002100",
                        "003951",
                        "002099"
                    ],
                    "Assegni Familiari": [
                        "ZP0160",
                        "ZP0162"
                    ]
                }
            }
        self.config = None

        self.__load_conversion_table()

    """ PRIVATE METHODS """
    def __load_conversion_table(self):
        """
        set PaycheckController configuration. with this configuration the program knows wich field
        extract from paychecks
        """
        if os.path.exists(self.conversion_table_path):
            with open(self.conversion_table_path, "r") as f:
                new_config = json.load(f)
                self.config = new_config
                return
        else:
            self.config = copy.deepcopy(self.default_configuration)


    """ PUBLIC METHODS """
    def create_config_from_csv(self, csv_path):
        """
        csv must have at least two columns 'Codice voce' and 'Descrizione voce'
        """
        new_config = copy.deepcopy(self.config)

        with open(csv_path, 'rt')as f:
            df = pd.read_csv(f, sep=";")
            columns = {}

            for index, row_ in df.iterrows():
                row = dict(row_)

                # parse out nan values
                parsed_row = {}
                for val in row:
                    if not pd.isnull(row[val]):
                        parsed_row[val] = row[val]

                if parsed_row['Codice voce'] not in new_config["col_codes"]:
                    columns[parsed_row['Codice voce']] = parsed_row
                    columns[parsed_row['Codice voce']]['col_value'] = -1

        # adding keys
        columns["quota t.f.r."] = {"Descrizione voce": "Quota T.F.R.", "col_value": -1}
        columns["quota t.f.r. a fondi"] = {"Descrizione voce": "Quota T.F.R. a Fondi", "col_value": -1}

        # setting new_config with parsed data and dumping it
        new_config["col_codes"] = columns
        with open(self.conversion_table_path, "w") as f:
            f.write(json.dumps(new_config, indent=4, ensure_ascii=True))

        self.__load_conversion_table()
        print(f"* * conversion_table.json created from this file >> {csv_path}")

    def validate_data(self):
        if not self.paychecks_to_check:
            raise Exception("Error: paycheck to check not specified")
        if not os.path.exists(self.paychecks_to_check):
            raise Exception(f"Error: cannot find {self.paychecks_to_check}")
        if not os.path.exists(self.badges_path):
            raise Exception(f"Error: cannot find {self.badges_path}")
        if not os.path.isdir(self.badges_path):
            raise Exception(f"Error: {self.badges_path} is not a folder")
        if len(os.listdir(self.badges_path)) < 1:
            raise Exception(f"Error: {self.badges_path} is empty!")

        return True

    # setters
    def set_badges_path(self, path):
        """setter for path to badges folder. it should contain .pdf files from every badge (splitted from BusinessCat)"""
        self.badges_path = path

    def set_paychecks_to_check_path(self, path):
        """setter for path to paychecks to check. it should be a .pdf file"""
        self.paychecks_to_check = path

    # main functions
    def paycheck_verification(self, create_Excel=True):
        PDF_file = self.paychecks_to_check
        total_content = {}

        col_codes = self.config['col_codes']
        col_to_extract = self.config['col_to_extract']
        merging_columns = self.config['merging_columns']

        # if there are no columns to extract set in the configuration, raise exception
        if not col_codes:
            raise Exception("No columns specified for extraction in config file")

        inputpdf = fitz.open(PDF_file)

        # per ogni pagina
        for i in range(inputpdf.pageCount):
            page = inputpdf.loadPage(i)
            blocks = page.getText("blocks")
            blocks.sort(key=lambda block: block[1])  # sort vertically ascending

            ####### find name in page
            name = None
            for index, b in enumerate(blocks):
                for elem in b:
                    elem = str(elem)
                    if ('cognome' in elem and 'nome' in elem) or ('COGNOME' in elem and 'NOME' in elem):
                        name_array = (blocks[index+1][4]).split('\n')

                        # conditional fix array
                        if len(name_array) == 2:
                            for check_ in name_array:
                                if "cessat" in check_.lower():
                                    name_array = (blocks[index+2][4]).split('\n')
                                    break

                        #check if there's a name in the array wich was found in the block
                        for value in name_array:
                            valid = True
                            for char in value:
                                if char.isdigit():
                                    valid = False
                                    break

                            if valid and value:
                                if value != None and not value.isspace() and value[0].isalpha():
                                    name = value
                                    break

            if name:
                name = name.upper()

                # add name to total_content
                if name not in total_content:
                    total_content[name] = {}

                ####### get page content
                for index, b in enumerate(blocks):
                    content = b[4].split('\n')
                    content = [string for string in content if (string != "" and not string.startswith("   "))]

                    # find ordinary and overtime hours in paycheck
                    h_check = [h.lower().strip() for h in content]
                    if "ore ordinarie" and "ore straordinarie" in h_check:
                        w_hours = [x for x in blocks[index+2][4].split('\n') if "," in x]
                        if len(w_hours) == 1:
                            w_hours.append("0")
                        try:
                            total_content[name]['Ore ordinarie'] = float(w_hours[0].replace(",","."))
                            total_content[name]['Ore straordinarie'] = float(w_hours[1].replace(",","."))
                        except IndexError:
                            pass

                    # parsing content
                    for i, elem in enumerate(content):

                        # parse netto del mese
                        if "NETTOsDELsMESE" in elem:
                            netto = blocks[index+1][-3]
                            netto = netto.replace("€", "").strip().replace(",",".")
                            netto = netto.replace(".", "", (netto.count('.') - 1))
                            try:
                                netto = float(netto)
                                total_content[name]["Netto del Mese"] = netto
                            except:
                                pass

                        elem = (elem.replace("*", "")).strip()

                        for paycheck_code in col_codes:

                            if (paycheck_code.lower() == elem.split(" ")[0].lower()) \
                                    or (len(paycheck_code.split(" ")) > 1) and elem.lower().startswith(paycheck_code.lower()):

                                # check if value is to be extracted
                                if paycheck_code in col_to_extract:

                                    # check if it needs to be merged
                                    elem_colname = None
                                    for colname in merging_columns:
                                        for subname in merging_columns[colname]:
                                            if subname == paycheck_code:
                                                elem_colname = colname

                                    if not elem_colname:
                                        elem_colname =  col_codes[paycheck_code]['Descrizione voce']

                                    # get row value
                                    val = content[col_codes[paycheck_code]['col_value']].replace(",", ".")
                                    val = val.replace(".", "", (val.count('.') - 1))
                                    try:
                                        val = float(val)
                                    except:
                                        pass

                                    # report data in the proper column
                                    if elem_colname not in total_content[name]:
                                        total_content[name][elem_colname] = val
                                    else:
                                        total_content[name][elem_colname] += val

        if create_Excel:
            sheet_name = "Verifica Buste Paga"
            self.create_Excel(total_content, sheet_name)
            print(f"File {self.verify_filename} generato con successo, {sheet_name} aggiunto al suo interno")

    def badges_verification(self, create_Excel=True):

        def parse_decimal_time(decimal_time):

            #if 0 not present attach it
            if len(decimal_time) == 1:
                decimal_time = decimal_time + "0"

            hour_min = 60
            result = (100*int(decimal_time))/hour_min
            return str(int(result))

        total_content = {}

        #for each badge
        for file in os.listdir(self.badges_path):
            path_to_badge = "/".join((self.badges_path, file))
            full_name = None
            last_block = []

            inputpdf = fitz.open(path_to_badge)

            ####### find name
            page = inputpdf.loadPage(0)
            blocks = page.getText("blocks")
            blocks.sort(key=lambda block: block[1])  # sort vertically ascending

            for index, b in enumerate(blocks):
                if "cognome e nome" in b[4].split('\n')[0].lower():
                    full_name = blocks[index+1][-3].split('\n')[0]
                    full_name = " ".join(full_name.split())
                    break

            ###### find values
            if full_name:
                total_content[full_name] = {"Ore ordinarie": 0.0, "Ore straordinarie": 0.0}

                for index, b in enumerate(blocks):

                    ##### check if block is a day
                    if len(b[4].split()[0][:-1]) <= 2 and b[4].split()[0][-1].isalpha() and b[4].split()[0][:-1].isdigit():
                        day = b[4].split()
                        day, day_values_ = day[0], day[1:]

                        # check if data is valid
                        if len(day[1:])%2 != 0 and day[1:][0].isdigit():
                            raise Exception(f"Expected even pairs, got uneven pairs for worker {full_name} on day {day}")

                        day_values_ = list(zip(day_values_[0::2], day_values_[1::2]))
                        day_values = []

                        # parse tuples in day values_
                        for i_, tupla in enumerate(day_values_):
                            # remove errors and start/end of workshift
                            if len((tupla[0])) > 4 or (len(tupla[0]) == 4 and "," not in tupla[0]):
                                continue
                            if len(day_values) < 2:
                                day_values.append(tupla)

                        # if not hours as first value skip day
                        if day_values:
                            if len(day_values[0][0]) == 3:
                                pass
                            else:
                                # parse decimal time values to number values
                                value_to_add = day_values[0][0].replace(",", ".").split(".")
                                if len(value_to_add) == 2:
                                    if value_to_add[1] != "00" and value_to_add[1] != "0":
                                        value_to_add[1] = parse_decimal_time(value_to_add[1])
                                    value_to_add = float(".".join(value_to_add))
                                else:
                                    value_to_add = float(day_values[0][0])
                                # if there's more than a pair of values in the day
                                if len(day_values) > 1:
                                    # if hours are overtime skip them
                                    if day_values[1][0] == "306" or day_values[1][0] == "006":
                                        pass
                                    else:
                                        total_content[full_name]['Ore ordinarie'] += value_to_add
                                # otherwise assume they are ordinary hours
                                else:
                                    total_content[full_name]['Ore ordinarie'] += value_to_add

                    ##### check if block is last and parse its data
                    if b[4].split()[0].isdigit() and not b[4].split()[0][-1].isalpha() and len(b[4].split()[0])<=3:
                        last_block_values = b[4].split()

                        # parse values
                        for val in last_block_values:
                            is_number = True
                            for letter in val:
                                if letter.isalpha() and letter != ",":
                                    is_number = False
                            # append conditionally
                            if is_number:
                                if "," in val:
                                    last_block.append(float(val.replace(",",".")))
                                else:
                                    last_block.append(int(val))
                            else:
                                if type(last_block[-1]) == int or type(last_block[-1]) == float:
                                    last_block.append(val)
                                else:
                                    last_block[-1] = last_block[-1] + " " + val

                        # check if data is valid
                        if len(last_block)%3 != 0:
                            raise Exception(f"Expected triplets on badge footer values for worker {full_name}")

                ##### last fixes
                total_content[full_name]["Ore ordinarie"] = total_content[full_name]["Ore ordinarie"]
                last_block = list(zip(last_block[0::3], last_block[1::3], last_block[2::3]))
                for pair in last_block:

                    #parse pair[2] time value to number value
                    pair_value = str(pair[2]).replace(",", ".").split(".")
                    if len(pair_value) == 2:
                        if pair_value[1] != "00" and pair_value[1] != "0":
                            pair_value[1] = parse_decimal_time(pair_value[1])
                        pair_value = float(".".join(pair_value))

                    if pair[0] == 306 or pair[0] == 6:
                        total_content[full_name]["Ore straordinarie"] += pair_value
                    total_content[full_name][str(pair[0]) + " " + str(pair[1])] = pair_value

        if create_Excel:
            sheet_name = "Verifica Cartellini"
            self.create_Excel(total_content, sheet_name)
            print(f"File {self.verify_filename} generato con successo, {sheet_name} aggiunto al suo interno")

    def create_Excel(self, content, sheet_name, transposed=True):
        df = pd.DataFrame.from_dict(content)

        if transposed:
            df = df.T

        # sort alphabetically rows and columns
        df = df.sort_index()
        df = df.reindex(sorted(df.columns), axis=1)

        open_mode = "a" if os.path.exists(self.verify_filename) else "w"

        with pd.ExcelWriter(self.verify_filename, mode=open_mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name)

    def compare_badges_to_paychecks(self, keep_refer_values=True):

        CHECK_SUFFIX = " PAYCHECK"
        badges_df = pd.read_excel(self.verify_filename, sheet_name="Verifica Cartellini", index_col=0)
        paychecks_df = pd.read_excel(self.verify_filename, sheet_name="Verifica Buste Paga", index_col=0)

        # set indexes name
        badges_df.index.name = "LAVORATORI"
        paychecks_df.index.name = "LAVORATORI"

        # fix wrong indexes in badges
        badges_df = badges_df.rename(index={
                                         'GUZMAN URENA ALEXANDER DE JE': 'GUZMAN URENA ALEXANDER DE JESUS',
                                         'NUTU LOREDANA ADRIAN': 'NUTU LOREDANA ADRIANA'
                                         })

        # uniform indexes
        badges_df.index = badges_df.index.str.upper()
        paychecks_df.index = paychecks_df.index.str.upper()

        # create df with all data
        common_columns = set(badges_df.columns.values).intersection(set(paychecks_df.columns.values))
        common_df = paychecks_df[list(common_columns)].copy()
        renaming = {key: key + CHECK_SUFFIX for key in common_df.columns.values}
        common_df = common_df.rename(columns=renaming)


        data_df = badges_df.merge(common_df, left_index=True, right_index=True)
        self.create_Excel(data_df, sheet_name="temp", transposed=False)

        destination_workbook = openpyxl.load_workbook(self.verify_filename)
        ws = destination_workbook["temp"]

        # find column of columns to highlight
        headings = [row for row in ws.iter_rows()][0]
        headings = [x.value for x in headings]
        matching_ = dict.fromkeys(common_columns, 0)
        matching = {}
        for col in matching_:
            matching[col] = 0
            matching[col + CHECK_SUFFIX] = 0
        for col in headings:
            if col in matching.keys():
                matching[col] = headings.index(col)

        for index, row in enumerate(ws.iter_rows()):
            # headers exclueded
            if index != 0:
                row_values = [x.value for x in row]
                worker_check = {}
                worker_errors = []

                # gather worker data
                for val in matching:
                    worker_check[val] = row_values[matching[val]]

                # find worker errors
                for data in worker_check:
                    check_val = data + CHECK_SUFFIX
                    if CHECK_SUFFIX not in data and check_val in worker_check:

                        if worker_check[data] == 0 and worker_check[check_val] == None:
                            continue
                        else:
                            if (worker_check[data]>0 and worker_check[check_val] == None) or (worker_check[data] - worker_check[check_val] != 0):
                                worker_errors.append(data)
                                worker_errors.append(check_val)

                if worker_errors:
                    # parse errors to cells
                    highlight_row = index + 1
                    for _i, error in enumerate(worker_errors):
                        highlight_column = get_column_letter(matching[error]+1)
                        worker_errors[_i] = str(highlight_column) + str(highlight_row)

                    for c in worker_errors:
                        cell = ws[c]
                        cell.fill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

        # drop refer columns conditionally
        if not keep_refer_values:
            col_to_remove = []
            for val in matching:
                if CHECK_SUFFIX in val:
                    col_to_remove.append(matching[val]+1)
            for val in sorted(col_to_remove, reverse=True):
                ws.delete_cols(val)


        #replace old verification with edited one
        destination_workbook.remove(destination_workbook["Verifica Cartellini"])
        ws.title = "Verifica Cartellini"
        destination_workbook.save(self.verify_filename)

        print(f">> BADGES COMPARED WITH PAYCHECKS SUCCESSFULLY")

        """
        # funziona male. evidenzia giusto ma non tutto
        styled_df = data_df.style
        for column in common_columns:
            col_to_check = column + " PAYCHECK"
            styled_df = styled_df.apply(lambda i_: ["background-color:red" if data_df.iloc[i_][col_to_check] - data_df.iloc[i_][column] != 0 else "" for i_, row in enumerate(data_df.iterrows())], subset=[column], axis=0)
        styled_df.to_excel("test.xlsx", engine="openpyxl", index=True)
        """

    def compare_paychecks_to_drive(self, df_bytestream, sheet, keep_refer_values=True, leave_blanks=False):

        CHECK_SUFFIX = " DRIVE"
        drive_df = get_comparison_df(df_bytestream, sheet)
        drive_df = drive_df[drive_df.index.notnull()]
        paychecks_df = pd.read_excel(self.verify_filename, sheet_name="Verifica Buste Paga", index_col=0)

        # set indexes name
        paychecks_df.index.name = "LAVORATORI"
        drive_df.index.name = "LAVORATORI"

        # uniform indexes
        drive_df.index = drive_df.index.str.upper()
        paychecks_df.index = paychecks_df.index.str.upper()

        # check divergences on dataframes
        problems = {"uncommon_indexes": [],"different_lenght": False, "error_string": ""}
        uncommon_indexes = list(set(drive_df.index.values) - set(paychecks_df.index.values))
        same_index_length = True if len(drive_df.index) - len(paychecks_df.index) == 0 else False
        # if there are not hired people in drive_df
        if (uncommon_indexes and not same_index_length):
            problems["uncommon_indexes"] = uncommon_indexes
            problems["error"] = "Non assunti sul Drive"
        # if there are typos in drive_df
        elif (uncommon_indexes and same_index_length):
            problems["uncommon_indexes"] = uncommon_indexes
            problems["error"] = "Errori di scrittura sul Drive"

        # merge dataframes
        common_columns = list(set(drive_df.columns.values).intersection(set(paychecks_df.columns.values)))
        common_df = drive_df[common_columns]
        common_df = common_df.rename(columns={key: key + CHECK_SUFFIX for key in common_df.columns.values})
        #data_df = pd.merge_ordered(left=common_df.reset_index(), right=paychecks_df.reset_index(), left_on="LAVORATORI", right_on="LAVORATORI", left_by="LAVORATORI").set_index("LAVORATORI").sort_index()
        data_df = paychecks_df.merge(common_df, left_index=True, right_index=True).sort_index()
        self.create_Excel(data_df, sheet_name="temp", transposed=False)


        destination_workbook = openpyxl.load_workbook(self.verify_filename)
        ws = destination_workbook["temp"]

        # create empty rows for uncommon_indexes (shouldnt be used)
        if leave_blanks:
            index_checkup = {k: v for k, v in enumerate(common_df.index.values)}
            # add blank lines based on index_checkup
            rows = list(enumerate(ws.iter_rows()))
            added_rows = 0
            for row in rows[::-1]:
                i_ = row[0]+1
                w_name = row[1][0].value
                try:
                    if index_checkup[i_-added_rows] != w_name:
                        ws.insert_rows(i_+added_rows)
                        added_rows += 1
                except:
                    pass

        # find column of columns to highlight
        headings = [row for row in ws.iter_rows()][0]
        headings = [x.value for x in headings]
        matching_ = dict.fromkeys(common_columns, 0)
        matching = {}
        for col in matching_:
            matching[col] = 0
            matching[col + CHECK_SUFFIX] = 0
        for col in headings:
            if col in matching.keys():
                matching[col] = headings.index(col)

        for index, row in enumerate(ws.iter_rows()):
            # headers exclueded
            if index != 0:
                row_values = [x.value for x in row]
                worker_check = {}
                worker_errors = []

                # gather worker data from his row
                for key in matching:
                    val = row_values[matching[key]]
                    if isinstance(val, str) and "€" in val:
                        val = val.replace("€", "").replace("-", "").replace(",", ".").strip()
                        worker_check[key] = float(val) if val else 0
                    elif not val:
                        worker_check[key] = 0
                    elif isinstance(val, float) or isinstance(val, int):
                        worker_check[key] = val

                # find worker errors
                for data in worker_check:
                    if CHECK_SUFFIX not in data:
                        try:
                            if worker_check[data] - worker_check[data + CHECK_SUFFIX] != 0:
                                worker_errors.append(data)
                                worker_errors.append(data + CHECK_SUFFIX)
                        except Exception as e:
                            print(e)

                if worker_errors:
                    # parse errors to cells
                    highlight_row = index + 1
                    for _i, error in enumerate(worker_errors):
                        highlight_column = get_column_letter(matching[error] + 1)
                        worker_errors[_i] = str(highlight_column) + str(highlight_row)

                    for c in worker_errors:
                        cell = ws[c]
                        cell.fill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

        # drop refer columns conditionally
        if not keep_refer_values:
            col_to_remove = []
            for val in matching:
                if CHECK_SUFFIX in val:
                    col_to_remove.append(matching[val] + 1)
            for val in sorted(col_to_remove, reverse=True):
                ws.delete_cols(val)

        # replace old verification with edited one
        destination_workbook.remove(destination_workbook["Verifica Buste Paga"])
        ws.title = "Verifica Buste Paga"
        destination_workbook.save(self.verify_filename)

        print(f">> PAYCHECKS COMPARED WITH DRIVE {sheet} VALUES SUCCESSFULLY")
        return problems

class BillingManager():
    def __init__(self, bill_name="Fattura"):
        self.bill_name = f"{bill_name}.xlsx"
        self.badges_path = None # badges_path
        self.regex_day_pattern = "([1-9]|[12]\d|3[01])[LMGVSF]"
        self.name_cell = "B5" # in che cella del badge_path si trova il nome nei cartellini
        self.pairing_schema = {
            "COD QTA": ["COD", "QTA"],
            "ENT USC": ["ENT", "USC"],
            "GIOR PROG": ["GIOR", "PROG"]
        }
        self.untouchable_keys = ["id", "tag"]
        self.total_content = None

        # model configs
        self.model_name = "Modello fatturazione.xlsx"
        self.footer_color = "e6e6e6"

        # config paths
        self._clients_path = "../config_files/BusinessCat billing/clients.json"
        self._billing_profiles_path = "../config_files/BusinessCat billing/billing_profiles.json"
        self._jobs_path = "../config_files/BusinessCat billing/jobs.json"

        # load data from config paths
        self.__load_clients()
        self.__load_billing_profiles()
        self.__load_jobs()

        # defaults
        self.default_new_job = {
            "id":"",
            "name":"",
            "billing_profile_id":""
        }
        self.default_new_client = {
            "id":"",
            "name":""
        }
        self.default_billing_profile = {
            "id": "",
            "name": "",
            "pricelist": [
                {
                    "tag": "OR",
                    "name": "ore_ordinarie",
                    "price": 0.0
                },
                {
                    "tag": "ST",
                    "name": "ore_straordinarie",
                    "price": 0.0
                },
                {
                    "tag": "MN",
                    "name": "ore_notturne",
                    "price": 0.0
                },
                {
                    "tag": "OF",
                    "name": "ore_festive",
                    "price": 0.0
                },
                {
                    "tag": "SF",
                    "name": "ore_straordinarie_festive",
                    "price": 0.0
                },
                {
                    "tag": "SN",
                    "name": "ore_straordinarie_notturne",
                    "price": 0.0
                },
                {
                    "tag": "FN",
                    "name": "ore_festive_notturne",
                    "price": 0.0
                }
            ]
        }

        print(">> BillingManager Initialized")


    """    PRIVATE METHODS    """
    def __get_engine(self):
        """get engine conditional based on extension of self.badges_path"""
        if not self.badges_path:
            raise Exception("badges_path missing!")
        elif self.badges_path.rsplit(".")[-1] == "xlsx":
            engine = "openpyxl"
        elif self.badges_path.rsplit(".")[-1] == "xls":
            engine = "xlrd"
        else:
            raise TypeError("self.badges_path is not an Excel!")
        return engine

    def __load_clients(self):
        """ read and load current billing_profiles file """

        # create empty file if not existing
        if not os.path.exists(self._clients_path):
            init_data = []
            with open(self._clients_path, "w") as f:
                f.write(json.dumps(init_data, indent=4, ensure_ascii=True))
            print("** created new clients.json file")

        with open(self._clients_path,"r") as f:
            self.clients = json.load(f)
        print("** clients caricati")

    def __load_billing_profiles(self):
        """ read and load current billing_profiles file """

        # create empty file if not existing
        if not os.path.exists(self._billing_profiles_path):
            init_data = []
            with open(self._billing_profiles_path, "w") as f:
                f.write(json.dumps(init_data, indent=4, ensure_ascii=True))
            print("** created new billing_profile.json file")

        with open(self._billing_profiles_path,"r") as f:
            self.billing_profiles = json.load(f)
        print("** billing_profiles caricati")

    def __load_jobs(self):
        """ read and load current billing_profiles file """

        # create empty file if not existing
        if not os.path.exists(self._jobs_path):
            init_data = []
            with open(self._jobs_path, "w") as f:
                f.write(json.dumps(init_data, indent=4, ensure_ascii=True))
            print("** created new jobs.json file")

        with open(self._jobs_path,"r") as f:
            self.jobs = json.load(f)
            self.jobs_namelist = sorted([job["name"] for job in self.jobs])
            self.jobs_namelist.insert(0,"")
        print("** jobs caricati")

    def __load_Excel_badges(self):
        """Load excel data of badges file (must be .xls or .xlsx)"""
        if not self.badges_path:
            raise Exception("ERROR: No badges_path specified")

        engine = self.__get_engine()
        try:
            if engine == "openpyxl":
                xlsx_data = openpyxl.load_workbook(self.badges_path)
                sheet_names = xlsx_data.sheetnames
            elif engine == "xlrd":
                xlsx_data = xlrd.open_workbook(self.badges_path, on_demand=True, logfile=open(os.devnull, 'w'))
                sheet_names = xlsx_data.sheet_names()
            else:
                raise
            return xlsx_data, sheet_names, engine
        except Exception as e:
            raise Exception(f"Cannot load_Excel_badges. Error: {e}")

    def __manage_columns(self, df):
        """ private method to fix original column names"""
        fixed_columns = []
        prev_fixed = False
        for index, v in enumerate(df.columns.values):
            if prev_fixed:
                prev_fixed = False
                continue
            new_value = df.columns.values[index].split()
            new_value = " ".join(new_value).strip()
            if new_value.startswith("COD QTA"):
                new_value = new_value.split()
                if len(new_value[1]) > 3:
                    new_value[0] = new_value[0] + new_value[1][3:]
                fixed_columns.append(new_value[0])
                fixed_columns.append(new_value[1])
                prev_fixed = True
            else:
                fixed_columns.append(new_value)
        df.columns = fixed_columns

        to_remove = []
        for index, col in enumerate(df.columns.values):
            if col.startswith("Unnamed"):
                # if col not in fixed_columns:
                to_remove.append(index)
        return df.drop(df.columns[to_remove], axis=1)

    def __get_badge_name(self, sheet_obj):
        """get owner's name out of sheet"""

        engine = self.__get_engine()
        try:
            if engine == "openpyxl":
                badge_name = (sheet_obj[self.name_cell]).value
            elif engine == "xlrd":
                badge_name = sheet_obj.cell_value(int(self.name_cell[1:]) - 1, int(openpyxl.utils.cell.column_index_from_string(self.name_cell[0])) - 1)
            else:
                raise
            try:
                badge_name = " ".join(badge_name.split())
            except:
                badge_name = None
            return badge_name
        except Exception as e:
            raise Exception(f"Cannot get_badge_name. Error: {e}")

    def __minutes_to_int(self, decimal_time):
        """ decimal_time => (str) MM"""
        # if 0 not present attach it
        if len(decimal_time) == 1:
            decimal_time = decimal_time + "0"

        hour_min = 60
        result = (100 * int(decimal_time)) / hour_min
        return str(int(result))

    def __round_float(self, float_number, decimal_pos=2):
        try:
            float(float_number)
        except:
            raise TypeError("Cannot round: not a float number")
        rounded = str(float_number).split(".")
        rounded = float(rounded[0] + "." + rounded[1][:decimal_pos])
        return rounded

    def __smart_renamer(self, name):
        old_name = name.split(" ")
        new_name = ""

        try:
            for index, word in enumerate(old_name):
                word = word.strip()
                if word:
                    if index < len(old_name):
                        new_name += (word[0].upper() + word[1:].lower() + " ")
                    else:
                        new_name += (word[0].upper() + word[1:].lower())
        except:
            new_name = name[0].upper() + name[1:].lower()

        return new_name

    def __gp_column_renamer(self, name):

        lookup_table = {
            "Ore ORD": "OR",
            "Ore STR": "ST",
            "Ore NOTT": "MN",
            "Ore FEST": "OF",
            "Ore STR/FEST": "SF",
            "Ore STR/NOTT": "SN",
            "Ore FEST/NOTT": "FN"
        }
        new_name = None
        if name in lookup_table:
            new_name = lookup_table[name]

        if not new_name:
            new_name = name

        return new_name

    def __apply_billing_profile(self, hours_to_bill, billing_profile):
        """
        steps: 1. adding time, 2. apply pattern, 3. apply pricing
        """
        priced_hours = {}

        # check integrity and get tag to focus
        if hours_to_bill["OR"] and hours_to_bill["OF"]:
            raise ValueError("ERROR: there are both ordinary and holiday hours on a single day")
        else:
            if hours_to_bill["OF"]:
                tag = "OF"
            elif hours_to_bill["OR"]:
                tag = "OR"
            else:
                tag = None

        # adding time (TIME IS ALWAYS ADDED IN BASE 100 eg. half an hour is not 0.30 but 0.5)
        if tag:
            if billing_profile["time_to_add"] and hours_to_bill[tag]:
                if billing_profile["add_over_threshold"]:
                    if hours_to_bill[tag] >= billing_profile["threshold_hour"]:
                        hours_to_bill[tag] += billing_profile["time_to_add"]
                else:
                    hours_to_bill[tag] += billing_profile["time_to_add"]

            # apply pattern
            if billing_profile["pattern"] and hours_to_bill[tag]:
                new_amount = 0.0
                start_val = copy.deepcopy(hours_to_bill[tag])
                for i in range(len(billing_profile["pattern"])):
                    operation = billing_profile["pattern"][i]["perform"].strip()
                    amount = billing_profile["pattern"][i]["amount"]
                    if operation == "/":
                        start_val /= amount
                    elif operation =="-":
                        start_val -= amount
                    elif operation == "+":
                        start_val += amount
                    elif operation =="*":
                        start_val *= amount
                    else:
                        raise Exception("ERROR: invalid operator in pattern. operator must be one of + - * /")
                    if billing_profile["pattern"][i]["keep"]:
                        new_amount += start_val
                hours_to_bill[tag] = new_amount

        # apply pricing
        try:
            if billing_profile["pricelist"]:
                for hour_type in hours_to_bill:
                    for p in billing_profile["pricelist"]:
                        if hour_type == p["tag"]:
                            priced_hours[hour_type] = hours_to_bill[hour_type]*p["price"]
                            priced_hours[hour_type] = self.__round_float(priced_hours[hour_type], decimal_pos=2)
            else:
                raise Exception("ERROR: No pricelist specified!")
        except:
            print("not found")

        return priced_hours

    def __get_gp_data(self, gp_xls):
        wb = openpyxl.load_workbook(gp_xls)
        ws = wb["Cartellino prova Query"]

        # prendo i nomi di colonna dalla prima riga
        columns = list(ws.iter_rows())[0]
        columns = [c.value for c in columns]
        columns = {k: v for k, v in enumerate(columns) if v}
        total_w = []
        w_obj = None
        for row in list(ws.iter_rows())[1:-1]:
            row_content = {k: v.value for k, v in enumerate(row) if k in columns}
            if row_content[1] and str(row_content[1]).lower().startswith("somma"): continue
            for x in row_content:
                if row_content[x] == None:
                    row_content[x] = 0

            if row_content[0]:
                if w_obj:
                    total_w.append(w_obj)
                w_obj = {}
                for item in columns.items():
                    w_obj[item[1]] = row_content[item[0]]
            else:
                for index in row_content:
                    key = columns[index]
                    if key != "Nome":
                        w_obj[key] += row_content[index]
        return total_w


    """     PROTECTED METHODS   """
    # must be called once before billing/creating model
    def _set_badges_path(self, badges_path):
        if not os.path.exists(badges_path):
            raise ValueError("ERROR: Cannot find badges path")
        self.badges_path = badges_path
        print("** badges_path caricato")

    # must be called once before billing/creating model
    def _set_billing_time(self, month, year):
        self.billing_year = int(year)
        self.billing_month = int(month)
        self._holidays = holidays.IT(years=[self.billing_year, self.billing_year - 1])

    def _parse_badges(self, names=[]):
        """
        read and fix the badges form, adjusting column names and preparing data to be read by other methods.
        returning a dict containing every worker as key and a subdict containing every of its workday as value
        """
        xlsx_data, sheet_names, engine = self.__load_Excel_badges()
        total_content = {}

        not_valid_names = "ENTE: DIV.: GRP/FILIALE: REP:. Cognome Nome Codice"
        not_valid_names = not_valid_names.split()

        for sheetNo, sheet in enumerate(sheet_names):
            sheet_data = xlsx_data[sheet] if engine == "openpyxl" else xlsx_data.get_sheet(sheetNo)

            valid_name = False
            incremented_by = 0
            self.name_cell = "B5"
            while not valid_name:
                badge_name = self.__get_badge_name(sheet_data)
                if badge_name and badge_name.split()[0] not in not_valid_names:
                    valid_name = True
                else:
                    incremented_by += 1
                    self.name_cell = f"B{int(self.name_cell[-1]) + 1}"
            if names:
                if not badge_name in names:
                    continue

            total_content[badge_name] = {}

            # getting df, fixing columns, removing empty columns
            df = pd.read_excel(xlsx_data, sheet_name=sheet, header=9 + incremented_by, index_col=0, engine=engine)

            # set columns
            df = self.__manage_columns(df)

            # parse rows
            for row_ in df.iterrows():
                i = str(row_[0]).strip()
                row = dict(row_[1])

                try:
                    if re.search(self.regex_day_pattern, i):
                        row_dict = {}

                        # pairing values
                        already_parsed = []
                        for val in row:
                            paired = False
                            for key in self.pairing_schema:
                                number = ""
                                if val.split(".")[0] in self.pairing_schema[key]:
                                    in_parsing = list(filter(lambda x: x != "", val.split(".")))
                                    if len(in_parsing) > 1 and in_parsing[1]:
                                        number = f".{in_parsing[1]}"

                                    main_key = key + "." + number if number else key + number
                                    if main_key not in already_parsed:
                                        row_dict[main_key] = []
                                        for v_ in self.pairing_schema[key]:

                                            # find the correct refer_key
                                            for _i in range(3):
                                                refer_key = v_ + "."*_i + number
                                                if refer_key in row.keys():
                                                    break

                                            if not isinstance(row[refer_key], pd.Series):
                                                row_dict[main_key].append(str(row[refer_key]).strip())
                                            else:
                                                check_val = ""
                                                for index in list(row[refer_key].to_list()):
                                                    if not str(index).isspace():
                                                        check_val = str(index).strip()
                                                row_dict[main_key].append(check_val)

                                        already_parsed.append(main_key)
                                    paired = True

                            if not paired:
                                row_dict[val] = row[val]

                        total_content[badge_name][i] = row_dict

                except TypeError:
                    pass

        self.total_content = total_content
        return total_content

    def _parse_days(self, total_content):
        """ Pointing out what is the type of the hours worked by the worker that day """

        to_return = {}
        for worker in total_content:
            to_return[worker] = {}
            for day in total_content[worker]:
                day_content = total_content[worker][day]
                parsed_day = {
                    "OR": 0.0,  # ordinarie
                    "ST": 0.0,  # straordinarie
                    "MN": 0.0,  # maggiorazione notturna
                    "OF": 0.0,  # ordinario festivo
                    "SF": 0.0,  # straordinario festivo
                    "SN": 0.0,  # straordinario notturno
                    "FN": 0.0  # festivo notturno
                }

                # se non ci sono ore ordinarie o straordinarie return empty day
                if not any(day_content["GIOR PROG"]) and not any(day_content["GIOR PROG..1"]):
                    to_return[worker][day] = parsed_day
                    continue

                # setting starting ordinary and overtime values
                if day_content["GIOR PROG"][0]:
                    val = day_content["GIOR PROG"][0] if len(day_content["GIOR PROG"][0]) >= 4 else "0" + day_content["GIOR PROG"][0]
                    if "." in val:
                        val = val.split(".")[0] + "." + self.__minutes_to_int(val.split(".")[1])
                    parsed_day["OR"] += float(val)
                if day_content["GIOR PROG..1"][0]:
                    val = day_content["GIOR PROG..1"][0] if len(day_content["GIOR PROG..1"][0]) >= 4 else "0" + day_content["GIOR PROG..1"][0]
                    if "." in val:
                        val = val.split(".")[0] + "." + self.__minutes_to_int(val.split(".")[1])
                    parsed_day["ST"] += float(val)

                # check every COD key for special hours
                for key in day_content:
                    if key.startswith("COD") and any(day_content[key]):

                        # night shifts
                        if day_content[key][0] == "MN":
                            hours = day_content[key][1] if len(day_content[key][1]) >= 4 else "0" + day_content[key][1]
                            if "." in hours:
                                hours = hours.split(".")[0] + "." + self.__minutes_to_int(hours.split(".")[1])
                            hours = float(hours)
                            parsed_day["OR"] -= hours
                            parsed_day["MN"] += hours

                        # overtime night shifts
                        elif day_content[key][0] == "SN":
                            hours = day_content[key][1] if len(day_content[key][1]) >= 4 else "0" + day_content[key][1]
                            if "." in hours:
                                hours = hours.split(".")[0] + "." + self.__minutes_to_int(hours.split(".")[1])
                            hours = float(hours)
                            parsed_day["ST"] -= hours
                            parsed_day["SN"] += hours

                # if day is holiday decrement ordinary to increase holiday values
                try:
                    check_day = f"{self.billing_month}/{day[:-1]}/{self.billing_year}"
                    if check_day in self._holidays:
                        parsed_day["OF"] += parsed_day["OR"]
                        parsed_day["OR"] -= parsed_day["OR"]
                        parsed_day["SF"] += parsed_day["ST"]
                        parsed_day["ST"] -= parsed_day["ST"]
                        parsed_day["FN"] += parsed_day["MN"]
                        parsed_day["MN"] -= parsed_day["MN"]
                except ValueError as e:
                    if str(e).startswith("Cannot parse date from string '2/29/"):
                        raise ValueError(f"ERRORE: {worker} ha lavorato il giorno 29 Febbraio di un anno non bisestile!")

                to_return[worker][day] = parsed_day
        return to_return

    def _new_job_id(self):
        id_lenght = 4

        check_high = 0
        for job in self.jobs:
            if int(job["id"]) > check_high:
                check_high = int(job["id"])
        check_high = str(check_high + 1) # increment 1 from the highest id found among all jobs

        new_id = "0"*(id_lenght-len(check_high)) + check_high
        return new_id

    def _new_client_id(self):
        id_lenght = 4

        check_high = 0
        for client in self.clients:
            if int(client["id"]) > check_high:
                check_high = int(client["id"])
        check_high = str(check_high + 1) # increment 1 from the highest id found among all jobs

        new_id = "0"*(id_lenght-len(check_high)) + check_high
        return new_id

    def _new_billing_profile_id(self):
        id_lenght = 4

        check_high = 0
        for profile in self.billing_profiles:
            if int(profile["id"]) > check_high:
                check_high = int(profile["id"])
        check_high = str(check_high + 1) # increment 1 from the highest id found among all jobs

        new_id = "0"*(id_lenght-len(check_high)) + check_high
        return new_id

    def _add_job(self):
        new_ = copy.deepcopy(self.default_new_job)
        new_["id"] = self._new_job_id()
        self.jobs.append(new_)
        return True

    def _rmv_job(self, pos):
        if len(self.jobs) > 0:
            self.jobs.pop(pos)
        return True

    def _add_client(self):
        new_ = copy.deepcopy(self.default_new_client)
        new_["id"] = self._new_client_id()
        self.clients.append(new_)
        return True

    def _rmv_client(self, pos):
        if len(self.clients) > 0:
            self.clients.pop(pos)
        return True

    def _add_billing_profile(self):
        new_ = copy.deepcopy(self.default_billing_profile)
        new_["id"] = self._new_billing_profile_id()
        self.billing_profiles.append(new_)
        return True

    def _rmv_billing_profile(self, pos):
        if len(self.billing_profiles) > 0:
            self.billing_profiles.pop(pos)
        return True

    def _create_model(self):
        sh_name = "Report Fatturazione"
        empty_rows_per_worker = 10
        footer_color = PatternFill(start_color=self.footer_color, end_color=self.footer_color, fill_type="solid")
        error_color = PatternFill(start_color=color_red[1:], end_color=color_red[1:], fill_type="solid")
        separator = Border(top=Side(border_style='thin', color="000000"))
        combobox_background = PatternFill(start_color="cce6ff", end_color="cce6ff", fill_type="solid")
        combobox_border = Border(bottom=Side(border_style='thin', color='e6f3ff'))
        combobox_font = openpyxl.styles.Font(bold=True, color="e67300")

        COLUMNS_TO_STYLE = ["J", "K", "L"] # style those columns with comboboxes parameters up here
        COLUMNS_TO_HIDE = {"first":"B","last":"H"}

        # parse data from given excel
        total_content = self._parse_badges()
        total_content = self._parse_days(total_content)
        totals = self.parse_total(total_content)
        total_workers_hours = totals[0]

        df = pd.DataFrame.from_dict(total_workers_hours).T

        # sort alphabetically rows and columns, renaming index
        df = df.sort_index()
        df.rename(index=lambda x: self.__smart_renamer(x), inplace=True)

        # adding "total" row and "total" column
        df.loc[">> ORE TOTALI <<"] = df.sum(axis=0, numeric_only=True)
        df['TOTALI'] = df.sum(axis=1, numeric_only=True)

        user_columns = [
            "cliente",
            "profilo",
            "mansione",
            "ore_ordinarie",
            "ore_straordinarie",
            "ore_notturne",
            "ore_festive",
            "ore_straordinarie_festive",
            "ore_straordinarie_notturne",
            "ore_festive_notturne",
            "totale"
        ]
        df_utente = pd.DataFrame(columns=user_columns)
        df = pd.concat([df,df_utente], axis=1)
        df.fillna("", inplace=True)

        df.index.rename("LAVORATORI", inplace=True)


        ############### GENERATING EXCEL MODEL
        with pd.ExcelWriter(self.model_name, mode="w") as writer:

            # create client, billing_profile, job lookup in a different sheet
            check_sheet = "check_sheet"
            check_ws = writer.book.create_sheet(check_sheet)
            check_ws.title = check_sheet
            vals_ = {
                "clients": [f"{x['id']} {x['name']}" for x in self.clients],
                "billing_profiles": [f"{x['id']} {x['name']}" for x in self.billing_profiles],
                "jobs": [f"{x['id']} {x['name']}" for x in self.jobs]
            }
            col_ = 1
            for key in vals_:
                row_ = 1
                for entry in vals_[key]:
                    check_ws[f"{get_column_letter(col_)}{row_}"].value = entry
                    row_ += 1
                col_ += 1

            # calculate check columns lenght
            check_sheet_columns = {
                "J":f'={check_sheet}!$A$1:$A$',
                "K":f'={check_sheet}!$B$1:$B$',
                "L":f'={check_sheet}!$C$1:$C$'
            }

            for i_, column in enumerate(check_ws.iter_cols(max_col=check_ws.max_column, max_row=check_ws.max_row)):
                referral = list(check_sheet_columns.keys())[i_]
                i_ += 1
                column_lenght = 0
                for cell in column:
                    if not cell.value:
                        break
                    column_lenght += 1
                if not column_lenght:
                    column_lenght += 1
                check_sheet_columns[referral] += str(column_lenght)

            # hide check_sheet
            check_ws.sheet_state = "hidden"

            ######################################################################################## MAIN BILL
            df.to_excel(writer, sheet_name=sh_name, na_rep=0, float_format="%.2f")
            ws = writer.sheets[sh_name]

            # style last rows
            last_rows_to_style = 1
            added_rows = 1 # fixed, don't touch
            for row in range(last_rows_to_style, 0, -1):
                for cell in ws[f"{(len(df) + 1 + added_rows) - row}:{(len(df) + 1 + added_rows) - row}"]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = footer_color

            # adding blank rows
            ws = writer.sheets[sh_name]
            refer_ws = copy.deepcopy(ws)
            check_name = None
            for row in refer_ws.iter_rows(max_row=ws.max_row):
                refer_name = row[0].value
                if refer_name and refer_name.upper() != refer_name and refer_name != check_name:
                    for index, row_ in enumerate(ws.iter_rows()):
                        if row_[0].value == refer_name:
                            ws.insert_rows(index+2, empty_rows_per_worker)
                            check_name = refer_name
                            break

            # set color, font weight, alignment of total column
            for cell in ws[f"I1:I{ws.max_row}"]:
                cell[0].font = openpyxl.styles.Font(bold=True)
                cell[0].fill = footer_color
                cell[0].alignment = Alignment(horizontal="center")
                cell[0].number_format = '#,##0.00'

            # adding formulas
            for index, row in enumerate(ws.iter_rows()):
                if index != 0 and index != 1 and index != ws.max_row:
                    row_total_formula = f"=SUM(M{index}:S{index})"
                    ws[f"T{index}"] = row_total_formula
                    ws[f"T{index}"].number_format = '#,##0.00'

                if row[0].value and row[0].value.upper() != row[0].value:
                    wtfi = index + (empty_rows_per_worker+1) # worker total formula index
                    worker_total_formula = f"=SUM(T{index+1}:T{wtfi})-I{index+1}"
                    ws[f"U{wtfi}"] = worker_total_formula
                    ws[f"U{wtfi}"].number_format = '#,##0.00'

            # adjust column width
            for index, row in enumerate(ws.iter_cols()):
                index +=1
                if index == 1:
                    ws.column_dimensions[get_column_letter(index)].width = 20
                elif index >=2 and index <=9:
                    ws.column_dimensions[get_column_letter(index)].width= 8
                elif index >= 10 and index <= 12:
                    ws.column_dimensions[get_column_letter(index)].width = 25
                else:
                    ws.column_dimensions[get_column_letter(index)].width = 15

            # adding comboboxes
            for col in COLUMNS_TO_STYLE:
                try:
                    dv = DataValidation(type="list", formula1=check_sheet_columns[col], allowBlank=True)
                    ws.add_data_validation(dv)
                    dv.add(f"{col}2:{col}{ws.max_row - 1}")
                except KeyError:
                    raise KeyError(f"ERRORE: nessuna lista di controllo per le celle della colonna {col}")
                for cell in ws[f"{col}2:{col}{ws.max_row-1}"]:
                    cell[0].font = combobox_font
                    cell[0].fill = combobox_background
                    cell[0].alignment = Alignment(horizontal="left")
                    cell[0].border = combobox_border

            # set border between workers
            for row in ws.iter_rows(max_row=ws.max_row):
                if row[0].value and row[0].value.upper() != row[0].value:
                    for cell in row:
                        # if already styled keep bottom style
                        if cell.border.bottom.color is not None or cell.border.bottom.style is not None:
                            temp_ = copy.deepcopy(separator)
                            temp_.bottom.color = cell.border.bottom.color
                            temp_.bottom.style = cell.border.bottom.style
                            cell.border = temp_
                            continue
                        cell.border = separator

            # hide columns
            ws.column_dimensions.group(COLUMNS_TO_HIDE["first"], COLUMNS_TO_HIDE["last"], outline_level=1, hidden=True)

            # freeze first column and row
            ws.freeze_panes = ws["B2"]

            # add conditional formatting
            ws.conditional_formatting.add(f'U2:U{ws.max_row}', formatting.rule.CellIsRule(operator='notEqual', formula=[0], fill=error_color))

    def _create_comparison(self, gp_filepath):
        sh_name = "Comparazione"
        footer_color = PatternFill(start_color=self.footer_color, end_color=self.footer_color, fill_type="solid")
        error_color = PatternFill(start_color=color_red[1:], end_color=color_red[1:], fill_type="solid")
        separator = Border(top=Side(border_style='thin', color="000000"))

        # parse data from given excel
        total_content = self._parse_badges()
        total_content = self._parse_days(total_content)
        totals = self.parse_total(total_content)
        total_workers_hours = totals[0]

        # get cartellini df
        df = pd.DataFrame.from_dict(total_workers_hours).T
        df.index.rename("LAVORATORI", inplace=True)
        df = df.sort_index()
        df.rename(index=lambda x: self.__smart_renamer(x), inplace=True)
        df['TOTALI'] = df.sum(axis=1, numeric_only=True)

        # get gp_df
        gp_data = self.__get_gp_data(gp_filepath)
        gp_df = pd.DataFrame.from_records(gp_data, index="Nome")
        gp_df = gp_df.sort_index()
        gp_df.rename(index=lambda x: self.__smart_renamer(x), inplace=True)
        gp_df.rename(columns=lambda x: self.__gp_column_renamer(x), inplace=True)
        gp_df['TOTALI'] = gp_df.sum(axis=1, numeric_only=True)
        gp_df.index.rename("LAVORATORI", inplace=True)

        # merge them
        df = pd.concat([df, gp_df], axis=1)
        df.index.rename("LAVORATORI", inplace=True)
        df.fillna(0, inplace=True)

        # adding "total" row
        df.loc[">> ORE TOTALI <<"] = df.astype(float).sum(axis=0, numeric_only=True)
        df = df.replace(0, "")


        ############### GENERATING EXCEL MODEL
        with pd.ExcelWriter("Comparazione.xlsx", mode="w") as writer:

            df.to_excel(writer, sheet_name=sh_name, na_rep=0, float_format="%.2f")
            ws = writer.sheets[sh_name]

            # style last rows
            last_rows_to_style = 1
            added_rows = 1 # fixed, don't touch
            for row in range(last_rows_to_style, 0, -1):
                for cell in ws[f"{(len(df) + 1 + added_rows) - row}:{(len(df) + 1 + added_rows) - row}"]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = footer_color

            # set color, font weight, alignment of totals column
            totals_columns = ["I", "R"]
            for col_ in totals_columns:
                for cell in ws[f"{col_}1:{col_}{ws.max_row}"]:
                    cell[0].font = openpyxl.styles.Font(bold=True)
                    cell[0].fill = footer_color
                    cell[0].alignment = Alignment(horizontal="center")
                    cell[0].number_format = '#,##0.00'

            # adding formulas
            for index, row in enumerate(ws.iter_rows()):
                if index != 0 and index != 1 and index != ws.max_row:
                    row_total_formula = f"=SUM(J{index}:Q{index})"
                    ws[f"R{index}"] = row_total_formula
                    ws[f"R{index}"].number_format = '#,##0.00'

                try:
                    if row[0].value and row[0].value.upper() != row[0].value:
                        worker_total_formula = f"=I{index}-R{index}"
                        ws[f"S{index}"] = worker_total_formula
                        ws[f"S{index}"].number_format = '#,##0.00'
                except AttributeError:
                    pass

            # adjust column width
            for index, row in enumerate(ws.iter_cols()):
                index +=1
                if index == 1:
                    ws.column_dimensions[get_column_letter(index)].width = 30
                elif (index >=2 and index <=8) or (index >=10 and index <=17):
                    ws.column_dimensions[get_column_letter(index)].width= 10
                else:
                    ws.column_dimensions[get_column_letter(index)].width = 15

            # set border between workers
            for row in ws.iter_rows(max_row=ws.max_row):
                try:
                    if row[0].value and row[0].value.upper() != row[0].value:
                        for cell in row:
                            # if already styled keep bottom style
                            if cell.border.bottom.color is not None or cell.border.bottom.style is not None:
                                temp_ = copy.deepcopy(separator)
                                temp_.bottom.color = cell.border.bottom.color
                                temp_.bottom.style = cell.border.bottom.style
                                cell.border = temp_
                                continue
                            cell.border = separator
                except AttributeError:
                    pass

            # freeze first column and row
            ws.freeze_panes = ws["B2"]

            # add conditional formatting
            ws["S1"].value = "DIFFERENZA"
            ws["S1"].font = openpyxl.styles.Font(bold=True)
            ws["S1"].alignment = Alignment(horizontal='center')
            ws.conditional_formatting.add(f'S2:S{ws.max_row}', formatting.rule.CellIsRule(operator='notEqual', formula=[0], fill=error_color))

    def _bill(self, model_path, profile_to_bill):

        BILLING_MODEL = openpyxl.load_workbook(model_path) # file to scan
        bpi = profile_to_bill # "<id> <name>"
        profile_obj = self.get_billing_profile_obj(bpi.split()[0]) # full_profile object
        pricelist = profile_obj["pricelist"]

        ### styles
        rows_between_jobs = 4
        job_font = openpyxl.styles.Font(bold=True, size=24)
        job_border = Border(bottom=Side(border_style='thick', color='3385ff'))
        header_font = openpyxl.styles.Font(bold=True)
        footer_color = openpyxl.styles.PatternFill(start_color=self.footer_color, end_color=self.footer_color, fill_type="solid")

        try:
            bill_name = f"Fattura {bpi.split()[1]} {self.billing_month}-{self.billing_year}.xlsx"
        except:
            bill_name = self.bill_name

        job_schema = ["client_id", "billing_profile_id", "job_id"]
        hours_type = ["OR", "ST", "MN", "OF", "SF", "SN", "FN"]

        try:
            ws = BILLING_MODEL["Report Fatturazione"]
        except:
            raise Exception("Non ho trovato il foglio 'Report Fatturazione' nel file excel!")

        # GROUP BY JOB
        grouped_by_job = {}
        active_name = None
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            name = row[0].value

            if name:
                #print(name)
                if name.upper() == name:
                    continue
                active_name = name
            if not name:
                name = active_name

            to_parse = [cell.value for cell in row[9:-2]] # get info from column J to column S
            job_info = to_parse[0:3] # cliente, profilo, mansione
            hours_info = to_parse[3:] # OR, ST, MN, OF, SF, SN, FN
            job_info = dict(zip(job_schema, job_info))
            hours_info = dict(zip(hours_type, hours_info))


            if job_info["billing_profile_id"] == bpi:

                if job_info["job_id"] not in grouped_by_job:
                    grouped_by_job[job_info["job_id"]] = {}

                if name not in grouped_by_job[job_info["job_id"]]:
                    grouped_by_job[job_info["job_id"]][name] = hours_info

        if not grouped_by_job:
            raise Exception("Il modello fornito non contiene lavoratori che abbiano svolto una mansione sotto quel profilo")

        # WRITE BILL
        title_ = bpi.split(" ", 1)[1]
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name("Sheet")) # remove default sheet
        wb.create_sheet(title_)
        wb.save(bill_name)

        row_ = 1
        for job_ in grouped_by_job:
            header_row = None
            ws = wb[title_]

            # get job df
            df = pd.DataFrame.from_dict(grouped_by_job[job_]).T
            df.index.rename("LAVORATORI", inplace=True)
            df.loc[">> ORE TOTALI <<"] = df.sum(axis=0, numeric_only=True)
            df['TOTALE'] = df.sum(axis=1, numeric_only=True)
            df.fillna(0.0, inplace=True)

            # write job name
            ws[f"A{row_}"].value = job_.split(" ", 1)[1]
            ws[f"A{row_}"].font = job_font
            for c_ in range(1,10):
                ws[f"{get_column_letter(c_)}{row_}"].border = job_border
            row_ += 1

            # save for reference
            header_row = row_

            # write headers
            col_ = 1
            ws[f"{get_column_letter(col_)}{header_row}"].value = df.index.name
            ws[f"{get_column_letter(col_)}{header_row}"].font = header_font
            col_ += 1
            for colname in df.columns.values:
                ws[f"{get_column_letter(col_)}{header_row}"].value = colname.upper()
                ws[f"{get_column_letter(col_)}{header_row}"].font = header_font
                col_ += 1
            row_ += 1

            # get lookup from headers
            lookup = {}
            for r in ws.iter_rows(min_row=header_row, max_row=header_row, max_col=ws.max_column):
                lookup = dict(zip([val.value for index, val in enumerate(r)],[index for index, val in enumerate(r)]))
                break

            # write workers
            for row in df.iterrows():
                col_ = 1
                worker = row[0]
                w_hours = dict(row[1])

                # per ogni intestazione inserisco il suo valore del lavoratore
                for h_type in lookup:
                    val_to_write = None
                    if h_type == "LAVORATORI":
                        val_to_write = worker
                    elif h_type == "TOTALE":
                        val_to_write = f"=SUM(B{row_}:H{row_})"
                    else:
                        try:
                            val_to_write = w_hours[h_type]
                        except KeyError:
                            print(f"key {w_hours[h_type]} not found!")

                    ws[f"{get_column_letter(col_)}{row_}"].value = val_to_write

                    # check style last row
                    if worker == ">> ORE TOTALI <<":
                        ws[f"{get_column_letter(col_)}{row_}"].fill = footer_color
                        ws[f"{get_column_letter(col_)}{row_}"].font = header_font

                    if col_ < ws.max_column:
                        col_ += 1

                # if last column and last row calculate billing hours for current job
                if worker == ">> ORE TOTALI <<":
                    billed_hours = dict(zip(lookup.keys(), [0.0 for entry in range(len(lookup.keys()))]))
                    for r in ws.iter_rows(min_row=header_row+1, max_row=row_, max_col=ws.max_column):
                        for h_type in lookup:
                            if h_type != "LAVORATORI":
                                if h_type != "TOTALE":
                                    colNo = lookup[h_type]
                                    to_sum = r[colNo].value
                                    billed_hours[h_type] += to_sum
                                else:
                                    billed_hours[h_type] = f"=SUM(B{row_+1}:H{row_+1})"

                    # move down a line
                    row_ += 1

                    # write billed hour row
                    for h_type in billed_hours:
                        colNo = lookup[h_type]
                        colNo += 1
                        if h_type == "LAVORATORI":
                            val_to_write = ">> IMPONIBILE <<"
                        else:

                            # multiply value for its price in pricelist
                            for p_ in pricelist:
                                if p_["tag"] == h_type:
                                    price_spec = p_
                                    val_to_write = billed_hours[h_type]
                                    val_to_write *= price_spec["price"]

                        try:
                            ws[f"{get_column_letter(colNo)}{row_}"].value = val_to_write
                            ws[f"{get_column_letter(colNo)}{row_}"].font = header_font
                            ws[f"{get_column_letter(colNo)}{row_}"].fill = footer_color
                            ws[f"{get_column_letter(colNo)}{row_}"].number_format = '#,##0.00€'
                        except:
                            raise Exception(f"ERRORE: valori imponibili non calcolabili")


                    row_ += rows_between_jobs
                else:
                    row_ += 1

        # adjust columns width
        for c_ in range(1,10):
            # first column
            if c_ == 1:
                ws.column_dimensions[get_column_letter(c_)].width = 40
            # next columns
            else:
                ws.column_dimensions[get_column_letter(c_)].width = 15
        wb.save(bill_name)

        return bill_name


    """    PUBLIC METHODS    """
    def get_all_badges_names(self):
        """ return an array of all names found in excel file """
        xlsx_data, sheet_names, engine = self.__load_Excel_badges()
        names = []
        for sheetNo, sheet in enumerate(sheet_names):
            sheet_data = xlsx_data[sheet] if engine == "openpyxl" else xlsx_data.get_sheet(sheetNo)
            names.append(self.__get_badge_name(sheet_data))
        return names

    def get_jobname(self, job_id):
        """ given id, gets job name """
        name = ""
        for job in self.jobs:
            if job["id"] == job_id:
                name = job["name"]
                break
        if not name and job_id:
            name = f"Job {job_id} non trovato"
        return name

    def get_client_object(self, client_id):
        client = None
        for c in self.clients:
            if c["id"] == client_id:
                client = c
                break
        return client

    def get_billing_profile_obj(self, billing_profile_id):
        billing_profile = None
        for profile in self.billing_profiles:
            if profile["id"] == billing_profile_id:
                billing_profile = profile
                break
        return billing_profile

    def get_billing_profile_id(self, job_id):
        """ return billing profile id given job id """
        billing_profile_id = ""
        for job in self.jobs:
            if job["id"] == job_id:
                billing_profile_id = job["billing_profile_id"]
                break
        if not billing_profile_id and job_id:
            raise Exception(f"Billing Profile for Job {job_id} non trovato")
        return billing_profile_id

    def parse_total(self, data, divided_by_job=False):
        """
        if divided_by_job == False return a tuple containing ({worker:total}, {total:total})
        if divided_by_job == True return a tuple containing ({job:{<worker>:total, job_total:job_total}}, {total:total})
        """
        total = {}
        new_data = {}

        if not divided_by_job:
            for worker in data:
                new_data[worker] = {}
                for day in data[worker]:
                    for hour_type in data[worker][day]:

                        # add to worker data
                        if hour_type in new_data[worker]:
                            new_data[worker][hour_type] += data[worker][day][hour_type]
                        else:
                            new_data[worker][hour_type] = data[worker][day][hour_type]

                        # add to total
                        if hour_type in total:
                            total[hour_type] += data[worker][day][hour_type]
                        else:
                            total[hour_type] = data[worker][day][hour_type]

                # round values
                for h in new_data[worker]:
                    new_data[worker][h] = self.__round_float(new_data[worker][h], decimal_pos=2)

        elif divided_by_job:
            for job in data:
                new_data[job] = {}
                for worker in data[job]:
                    new_data[job][worker] = {}
                    for day in data[job][worker]:
                        for hour_type in data[job][worker][day]:
                            if hour_type not in new_data[job][worker]:
                                new_data[job][worker][hour_type] = 0.0
                            new_data[job][worker][hour_type] += data[job][worker][day][hour_type]

                # parse total for job and round values
                job_total = {}
                for worker_ in new_data[job]:
                    for hour_type in new_data[job][worker_]:

                        # add to job_total
                        if hour_type not in job_total:
                            job_total[hour_type] = 0.0
                        job_total[hour_type] += new_data[job][worker_][hour_type]

                        # add to total
                        if hour_type not in total:
                            total[hour_type] = 0.0
                        total[hour_type] += new_data[job][worker_][hour_type]

                        new_data[job][worker_][hour_type] = self.__round_float(new_data[job][worker_][hour_type],
                                                                               decimal_pos=2)
                new_data[job]["job_total"] = job_total

        # round values
        for h in total:
            total[h] = self.__round_float(total[h], decimal_pos=2)

        return (new_data, total)











    '''

    # probably will be deleted
    def parse_jobs_to_profiles(self, workers_jobs):
        """ creating a dict from parsing every worker day to its billing profile and returning it """
        workers_billing_profiles = {}
        for w in workers_jobs:
            workers_billing_profiles[w] = {}
            for day in workers_jobs[w]:
                if not workers_jobs[w][day]:
                    workers_billing_profiles[w][day] = ""
                else:
                    workers_billing_profiles[w][day] = self.get_billing_profile_id(workers_jobs[w][day])
        return workers_billing_profiles

    def _bill(self, hours, jobs, billing_profiles, bill_by_job=True, dump_values=False, dump_detailed=False):
        billed_hours = {}

        ####### unic sheet
        if not bill_by_job:
            for w in hours:
                billed_hours[w] = {}
                w_hours = hours[w]
                w_jobs = jobs[w]
                w_billing_profiles = billing_profiles[w]

                for day in w_hours:
                    day_job = w_jobs[day]
                    day_billing_profile_id = w_billing_profiles[day]
                    day_billing_profile = self.get_billing_profile_obj(day_billing_profile_id)

                    # if worker worked that day bill it, else append 0 values
                    if day_job:
                        billed_hours[w][day] = self.__apply_billing_profile(w_hours[day], day_billing_profile)
                    else:
                        billed_hours[w][day] = w_hours[day]

            new_billed_hours, total_billing = self.parse_total(billed_hours, divided_by_job=False)
            new_hours_data, total_hours = self.parse_total(hours, divided_by_job=False)
            self.create_Excel(new_hours_data, total_billing, bill_by_job=bill_by_job)

        #### a sheet for every job
        elif bill_by_job:
            hours_by_job = {}
            for w in jobs:
                if w != "job_total":
                    for day in jobs[w]:
                        current_job = jobs[w][day]
                        if current_job:
                            current_job = self.get_jobname(current_job)
                            if current_job not in hours_by_job:
                                hours_by_job[current_job] = {}
                            if w not in hours_by_job[current_job]:
                                hours_by_job[current_job][w] = {}
                            hours_by_job[current_job][w][day] = hours[w][day]

            for job in hours_by_job:
                billed_hours[job] = {}
                for w in hours_by_job[job]:
                    if w != "job_total":
                        billed_hours[job][w] = {}
                        for day in hours_by_job[job][w]:
                            billed_hours[job][w][day] = self.__apply_billing_profile(hours_by_job[job][w][day],self.get_billing_profile_obj(billing_profiles[w][day]))

            new_billed_hours, total_billing = self.parse_total(billed_hours, divided_by_job=True)
            new_hours_data, total_hours = self.parse_total(hours_by_job, divided_by_job=True)
            self.create_Excel(new_hours_data, new_billed_hours, bill_by_job=bill_by_job)

        # conditional dump values
        if dump_detailed:
            with open("DETAIL_ore_lavoratori.json", "w") as f:
                f.write(json.dumps(hours, indent=4, ensure_ascii=True))
            with open("DETAIL_valori_da_fatturare.json", "w") as f:
                f.write(json.dumps(billed_hours, indent=4, ensure_ascii=True))

        if dump_values:
            with open("ore_lavoratori.json", "w") as f:
                f.write(json.dumps(new_hours_data, indent=4, ensure_ascii=True))
            with open("valori_da_fatturare.json", "w") as f:
                f.write(json.dumps(new_billed_hours, indent=4, ensure_ascii=True))

        print(">> Billed Successfully")
        
    def create_Excel(self, data, total_billing, transposed=True, bill_by_job=False):

        for job in data:

            total_ = data[job].pop("job_total") if bill_by_job else total_billing

            # adjust to billing type
            if bill_by_job:
                df = pd.DataFrame.from_dict(data[job])
            else:
                job = "Report Fatturazione"
                df = pd.DataFrame.from_dict(data)

            if transposed:
                df = df.T

            # sort alphabetically rows and columns
            df = df.sort_index()
            df.rename(index=lambda x: self.__smart_renamer(x), inplace=True)

            # add totals (rows total/column total)
            df.loc[">> ORE TOTALI <<"] = df.sum(axis=0, numeric_only=True)
            if bill_by_job:
                df.loc[">> € DA FATTURARE <<"] = total_
            df['TOTALI'] = df.sum(axis=1, numeric_only=True)

            # polish
            df.index.rename("LAVORATORI", inplace=True)
            df.replace(0, np.nan, inplace=True)

            #generating excel with df data
            mode = "a" if os.path.exists(self.bill_name) else "w"
            with pd.ExcelWriter(self.bill_name, mode=mode) as writer:
                df.to_excel(writer, sheet_name=job, na_rep=0, float_format="%.2f")
                ws = writer.sheets[job]
                footer_color = PatternFill(start_color="e6e6e6", end_color="e6e6e6", fill_type="solid")

                # color last rows
                last_rows_to_style = 2 if bill_by_job else 1
                added_rows = 2 if bill_by_job else 1
                for row in range(last_rows_to_style, 0, -1):
                    for cell in ws[f"{(len(df)+1+added_rows)-row}:{(len(df)+1+added_rows)-row}"]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = footer_color

                        if bill_by_job and row == last_rows_to_style:
                            cell.number_format = '#,##0.00€'

            # if single sheet result break
            if not bill_by_job:
                break
    
    '''